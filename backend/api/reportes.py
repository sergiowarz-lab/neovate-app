import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.api.deps import get_current_user, require_admin_or_analista
from backend.core.database import get_db
from backend.models import (
    EstadoReporte, EstadoSeguimiento, ReporteNomina, ReporteSS,
    RolUsuario, SeguimientoMensual, Usuario,
)
from backend.schemas.reporte import ReporteNominaOut, ReporteSSOut


router = APIRouter(
    prefix="/api/reportes",
    tags=["reportes"],
    dependencies=[Depends(get_current_user)],
)


def _nit9_forzado(current_user: Usuario, nit9: str | None) -> str | None:
    """Si el usuario es rol empresa, fuerza el filtro a su propio nit_empresa."""
    if current_user.rol == RolUsuario.EMPRESA:
        return current_user.nit_empresa
    return nit9


@router.get("/ss", response_model=list[ReporteSSOut])
def historial_ss(
    nit9: str | None = None,
    anio: int | None = None,
    mes: int | None = None,
    estado: EstadoReporte | None = None,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[ReporteSS]:
    nit9_q = _nit9_forzado(current_user, nit9)
    stmt = select(ReporteSS)
    if nit9_q:
        stmt = stmt.where(ReporteSS.nit9 == nit9_q)
    if anio is not None:
        stmt = stmt.where(ReporteSS.anio == anio)
    if mes is not None:
        stmt = stmt.where(ReporteSS.mes_obligacion == mes)
    if estado:
        stmt = stmt.where(ReporteSS.estado == estado)
    stmt = stmt.order_by(ReporteSS.anio.desc(), ReporteSS.mes_obligacion.desc(), ReporteSS.id.desc())
    return list(db.scalars(stmt))


@router.get("/ss/{reporte_id}", response_model=ReporteSSOut)
def detalle_ss(
    reporte_id: str,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ReporteSS:
    rep = db.get(ReporteSS, reporte_id)
    if not rep:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Reporte SS no encontrado")
    if current_user.rol == RolUsuario.EMPRESA and rep.nit9 != current_user.nit_empresa:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Sin acceso a este reporte")
    return rep


@router.get("/nomina", response_model=list[ReporteNominaOut])
def historial_nomina(
    nit9: str | None = None,
    anio: int | None = None,
    mes: int | None = None,
    estado: EstadoReporte | None = None,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[ReporteNomina]:
    nit9_q = _nit9_forzado(current_user, nit9)
    stmt = select(ReporteNomina)
    if nit9_q:
        stmt = stmt.where(ReporteNomina.nit9 == nit9_q)
    if anio is not None:
        stmt = stmt.where(ReporteNomina.anio == anio)
    if mes is not None:
        stmt = stmt.where(ReporteNomina.mes_obligacion == mes)
    if estado:
        stmt = stmt.where(ReporteNomina.estado == estado)
    stmt = stmt.order_by(ReporteNomina.anio.desc(), ReporteNomina.mes_obligacion.desc())
    return list(db.scalars(stmt))


@router.get("/nomina/{reporte_id}", response_model=ReporteNominaOut)
def detalle_nomina(
    reporte_id: str,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ReporteNomina:
    rep = db.get(ReporteNomina, reporte_id)
    if not rep:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Reporte Nómina no encontrado")
    if current_user.rol == RolUsuario.EMPRESA and rep.nit9 != current_user.nit_empresa:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Sin acceso a este reporte")
    return rep


# ─── Exportaciones Excel (solo admin y analista) ───────────────────────────────

@router.get("/exportar/resumen")
def exportar_resumen(
    anio: int,
    mes: int,
    _: Usuario = Depends(require_admin_or_analista),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Excel con el estado SS de todas las empresas para el período dado."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    rows = list(db.scalars(
        select(SeguimientoMensual)
        .where(SeguimientoMensual.anio == anio, SeguimientoMensual.mes == mes)
        .order_by(SeguimientoMensual.nombre_empresa)
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Resumen {mes:02d}-{anio}"

    header_fill = PatternFill("solid", fgColor="1A1825")
    header_font = Font(bold=True, color="C9A84C")
    headers = [
        "NIT9", "Empresa", "Estado SS", "Fecha Plazo SS", "Fecha Pago SS",
        "Días Mora SS", "Estado Nómina", "Fecha Plazo Nómina", "Fecha Pago Nómina",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    estado_colors = {
        "Cumple": "2ECC71", "En_mora": "E74C3C", "Pendiente": "F39C12",
    }
    for row_idx, seg in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=seg.nit9)
        ws.cell(row=row_idx, column=2, value=seg.nombre_empresa)
        ws.cell(row=row_idx, column=3, value=seg.ss_estado.value if seg.ss_estado else "")
        ws.cell(row=row_idx, column=4, value=str(seg.ss_fecha_plazo) if seg.ss_fecha_plazo else "")
        ws.cell(row=row_idx, column=5, value=str(seg.ss_fecha_pago) if seg.ss_fecha_pago else "")
        ws.cell(row=row_idx, column=6, value=seg.ss_dias_mora or 0)
        ws.cell(row=row_idx, column=7, value=seg.nomina_estado.value if seg.nomina_estado else "")
        ws.cell(row=row_idx, column=8, value=str(seg.nomina_fecha_plazo) if seg.nomina_fecha_plazo else "")
        ws.cell(row=row_idx, column=9, value=str(seg.nomina_fecha_pago) if seg.nomina_fecha_pago else "")

        color = estado_colors.get(seg.ss_estado.value if seg.ss_estado else "", "FFFFFF")
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=color + "33")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"neovate_resumen_{anio}_{mes:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/exportar/mora")
def exportar_mora(
    anio: int,
    mes: int,
    _: Usuario = Depends(require_admin_or_analista),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Excel solo con empresas en mora para el período dado."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    rows = list(db.scalars(
        select(SeguimientoMensual)
        .where(
            SeguimientoMensual.anio == anio,
            SeguimientoMensual.mes == mes,
            SeguimientoMensual.ss_estado == EstadoSeguimiento.EN_MORA,
        )
        .order_by(SeguimientoMensual.ss_dias_mora.desc())
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mora {mes:02d}-{anio}"

    header_fill = PatternFill("solid", fgColor="C0392B")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["NIT9", "Empresa", "Días en Mora", "Fecha Plazo SS", "Fecha Pago SS"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, seg in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=seg.nit9)
        ws.cell(row=row_idx, column=2, value=seg.nombre_empresa)
        ws.cell(row=row_idx, column=3, value=seg.ss_dias_mora or 0)
        ws.cell(row=row_idx, column=4, value=str(seg.ss_fecha_plazo) if seg.ss_fecha_plazo else "")
        ws.cell(row=row_idx, column=5, value=str(seg.ss_fecha_pago) if seg.ss_fecha_pago else "")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"neovate_mora_{anio}_{mes:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
